import openpyxl
from openpyxl.styles import Font, PatternFill
import os
from xml.etree import ElementTree as ET

SCREEN_WIDTH = 90


class Weapon:
    def __init__(self, name, rate_of_fire, shots_in_burst, release_time_after_fire, deviate_shot_angle, ammo_magazine_subtype, reload_time, ammo_name, desired_speed, max_trajectory, is_explosive, backkick_force,
                 physical_material, projectile_hit_impulse, missile_explosion_radius, projectile_mass_damage,
                 missile_explosion_damage, projectile_health_damage, missile_health_damage, ammo_magazine_name, ammo_magazine_capacity):
        # (String) Used for storing the name of the weapon
        self.name = name
        # (Integer) Used for storing how many shots per second the weapon will fire
        self.rate_of_fire = rate_of_fire
        # (Integer) Used for storing how many shots the weapon will fire in a burst (default 1)
        self.shots_in_burst = shots_in_burst
        # (Integer) Used for storing how many milliseconds the weapon will wait after firing
        self.release_time_after_fire = release_time_after_fire
        # (Float) Used for storing the angle (in degrees?) that the shot may be off by
        self.deviate_shot_angle = deviate_shot_angle
        # (String) Used for storing the name of the ammo magazine for this weapon
        self.ammo_magazine_subtype = ammo_magazine_subtype
        # (Integer) Used for storing the time it takes to reload the ammo magazine
        self.reload_time = reload_time
        # (String) Used for storing the name of the ammo
        self.ammo_name = ammo_name
        # (Integer) Used for storing the speed of the projectile as it travels towards its target
        self.desired_speed = desired_speed
        # (Integer) Used for storing the range of the projectile before disappearing or exploding
        self.max_trajectory = max_trajectory
        # (String) Used for storing whether the projectile deals explosive damage or doesn't
        self.is_explosive = is_explosive
        # (Integer) Used for storing the recoil value of the weapon when firing
        self.backkick_force = backkick_force
        # (String) Used for storing the type of projectile this is
        self.physical_material = physical_material
        # (Integer) Used for storing the physical impact value on the target
        self.projectile_hit_impulse = projectile_hit_impulse
        # (Integer) Used for storing the number of meters the explosion will effect from its detonation point
        self.missile_explosion_radius = missile_explosion_radius
        # (Integer) Used for storing the damage value of the projectile
        self.projectile_mass_damage = projectile_mass_damage
        # (Integer) Used for storing the damage value of the projectile
        self.missile_explosion_damage = missile_explosion_damage
        # (Integer) Used for storing the damage applied to the health of nearby players
        self.projectile_health_damage = projectile_health_damage
        # (Integer) Used for storing the damage applied to the health of nearby players
        self.missile_health_damage = missile_health_damage
        # (String) Used for storing the name of the ammo magazine
        self.ammo_magazine_name = ammo_magazine_name
        # (Integer) Used for storing the amount of ammo stored inside this magazine item
        self.ammo_magazine_capacity = ammo_magazine_capacity


class Ammo:
    def __init__(self, name, desired_speed, max_trajectory, is_explosive, backkick_force,
                 physical_material, projectile_hit_impulse, missile_explosion_radius, projectile_mass_damage,
                 missile_explosion_damage, projectile_health_damage, missile_health_damage):
        self.name = name
        # (Integer) Used for storing the speed of the projectile as it travels towards its target
        self.desired_speed = desired_speed
        # (Integer) Used for storing the range of the projectile before disappearing or exploding
        self.max_trajectory = max_trajectory
        # (String) Used for storing whether the projectile deals explosive damage or doesn't
        self.is_explosive = is_explosive
        # (Integer) Used for storing the recoil value of the weapon when firing
        self.backkick_force = backkick_force
        # (String) Used for storing the type of projectile this is
        self.physical_material = physical_material
        # (Integer) Used for storing the physical impact value on the target
        self.projectile_hit_impulse = projectile_hit_impulse
        # (Integer) Used for storing the number of meters the explosion will effect from its detonation point
        self.missile_explosion_radius = missile_explosion_radius
        # (Integer) Used for storing the damage value of the projectile
        self.projectile_mass_damage = projectile_mass_damage
        # (Integer) Used for storing the damage value of the projectile
        self.missile_explosion_damage = missile_explosion_damage
        # (Integer) Used for storing the damage applied to the health of nearby players
        self.projectile_health_damage = projectile_health_damage
        # (Integer) Used for storing the damage applied to the health of nearby players
        self.missile_health_damage = missile_health_damage
        # (String) Used for storing the name of the ammo magazine


class AmmoMagazine:
    def __init__(self, name, capacity):
        # (String) Used for storing the name of the ammo magazine
        self.name = name
        # (Integer) Used for storing the amount of ammo stored inside this magazine item
        self.capacity = capacity


WEAPON_KEY_STRING_PHRASES = ['<SubtypeId>', '<AmmoMagazine Subtype=']
WEAPON_KEY_INT_PHRASES = ['RateOfFire=', 'ShotsInBurst=', '<ReleaseTimeAfterFire>', '<ReloadTime>']
WEAPON_KEY_FLOAT_PHRASES = ['<DeviateShotAngle>']
AMMO_KEY_STRING_PHRASES = ['<SubtypeId>', '<PhysicalMaterial>', '<IsExplosive>']
AMMO_KEY_INT_PHRASES = ['<DesiredSpeed>', '<MaxTrajectory>', '<BackkickForce>', '<ProjectileHitImpulse>', '<MissileExplosionRadius>', '<ProjectileMassDamage>', '<MissileExplosionDamage>', '<ProjectileHealthDamage>', '<MissileHealthDamage>']
AMMO_MAGAZINE_KEY_STRING_PHRASES = ['<SubtypeId>']
AMMO_MAGAZINE_KEY_INT_PHRASES = ['<Capacity>']


# Prompt the user to select a data file to be parsed
# Create a list of all the files in the CWD that could be parsed and assign them a number
def main():
    local_weapon_list = []
    local_ammo_magazine_list = []
    local_ammo_list = []
    target_file = 1
    local_cwd = os.getcwd()
    local_cwd_files = {}
    file_number = 0
    for file in reversed(os.listdir(local_cwd)):
        if file.endswith('.sbc'):
            file_number += 1
            local_cwd_files.update({file_number: file})
            directory_quantity = len(local_cwd_files)
        else:
            pass
    while True:
        print("-" * SCREEN_WIDTH)
        for key, value in local_cwd_files.items():
            print(key, value)
        print("-" * SCREEN_WIDTH)
        print("Press Enter to parse all .sbc files.")
        print("Just close this program from this menu when you're done.")
        garbage_input = input()
        break
    while target_file <= directory_quantity:
        # os.system('cls')
        # We now have a spreadsheet with the headers established and frozen, time to read the data.
        print("Working")
        local_file = open(local_cwd_files[int(target_file)])
        print("Processing file {}".format(str(local_file)))
        local_current_object = None
        local_whitespace_count = 0
        while True:
            local_line = local_file.readline()
            # Check the lines contents
            if not local_line:
                # There's nothing inside
                if local_current_object:
                    local_current_object = None
                local_whitespace_count += 1
                if local_whitespace_count > 25:
                    # The file must be out of content
                    break
                pass
            elif '<Weapon>' in local_line:
                # We have encountered the start of a weapon description
                local_current_object = Weapon('n/a', 0, 1, 0, 0.0, 'n/a', 0, 'n/a', 0, 0, 'n/a', 0, 'n/a', 0, 0, 0, 0, 0, 0, 'n/a', 0)
                while True:
                    local_line = local_file.readline()
                    if not local_line:
                        # There's nothing inside
                        local_current_object = None
                        break
                    else:
                        if "</Weapon>" in local_line:
                            # The object is complete
                            print("Weapon object created")
                            print("Object named: {}".format(local_current_object.name))
                            local_weapon_list.append(local_current_object)
                            local_current_object = None
                            break
                        for phrase in WEAPON_KEY_STRING_PHRASES:
                            # Check to see if any phrases are in the string
                            if phrase in local_line:
                                # The phrase is in the string
                                # Remove the value from the string
                                local_current_value = str(remove_term(local_line, phrase))
                                # Apply the value to the corresponding object attribute
                                update_attr(local_current_object, phrase, local_current_value)
                            else:
                                pass
                        for phrase in WEAPON_KEY_FLOAT_PHRASES:
                            if phrase in local_line:
                                # The phrase is in the string
                                # Remove the value from the string
                                local_current_value = float(remove_term(local_line, phrase))
                                # Apply the value to the corresponding object attribute
                                update_attr(local_current_object, phrase, local_current_value)
                            else:
                                pass
                        for phrase in WEAPON_KEY_INT_PHRASES:
                            if phrase in local_line:
                                # The phrase is in the string
                                # Remove the value from the string
                                local_current_value = int(remove_term(local_line, phrase))
                                # Apply the value to the corresponding object attribute
                                update_attr(local_current_object, phrase, local_current_value)
                            else:
                                pass
            elif '<Ammo xsi:type=' in local_line:
                # We have encountered the start of an ammo description
                local_current_object = Ammo('n/a', 0, 0, 'n/a', 0, 'n/a', 0, 0, 0, 0, 0, 0)
                while True:
                    local_line = local_file.readline()
                    if not local_line:
                        # There's nothing inside
                        local_current_object = None
                        break
                    else:
                        if "</Ammo>" in local_line:
                            # The object is complete
                            local_ammo_list.append(local_current_object)
                            print("Ammo object created")
                            print("Object named: {}".format(local_current_object.name))
                            local_current_object = None
                            break
                        for phrase in AMMO_KEY_STRING_PHRASES:
                            # Check to see if any phrases are in the string
                            if phrase in local_line:
                                # The phrase is in the string
                                # Remove the value from the string
                                local_current_value = str(remove_term(local_line, phrase))
                                # Apply the value to the corresponding object attribute
                                update_attr(local_current_object, phrase, local_current_value)
                            else:
                                pass
                        for phrase in AMMO_KEY_INT_PHRASES:
                            if phrase in local_line:
                                # The phrase is in the string
                                # Remove the value from the string
                                local_current_value = (int(float(remove_term(local_line, phrase))))
                                # Apply the value to the corresponding object attribute
                                update_attr(local_current_object, phrase, local_current_value)
                            else:
                                pass
            elif '<AmmoMagazine>' in local_line:
                # We have encountered the start of an ammo magazine description
                local_current_object = AmmoMagazine('n/a', 0)
                while True:
                    local_line = local_file.readline()
                    if not local_line:
                        # There's nothing inside
                        local_current_object = None
                        break
                    else:
                        if "</AmmoMagazine>" in local_line:
                            # The object is complete
                            local_ammo_magazine_list.append(local_current_object)
                            print("Ammo Magazine Object created")
                            print("Object named: {}".format(local_current_object.name))
                            local_current_object = None
                            break
                        for phrase in AMMO_MAGAZINE_KEY_STRING_PHRASES:
                            # Check to see if any phrases are in the string
                            if phrase in local_line:
                                # The phrase is in the string
                                # Remove the value from the string
                                local_current_value = str(remove_term(local_line, phrase))
                                # Apply the value to the corresponding object attribute
                                update_attr(local_current_object, phrase, local_current_value)
                            else:
                                pass
                        for phrase in AMMO_MAGAZINE_KEY_INT_PHRASES:
                            if phrase in local_line:
                                # The phrase is in the string
                                # Remove the value from the string
                                local_current_value = int(remove_term(local_line, phrase))
                                # Apply the value to the corresponding object attribute
                                update_attr(local_current_object, phrase, local_current_value)
                            else:
                                pass
            elif '</Definitions>' in local_line:
                # The file is out of data
                target_file += 1
                break
            else:
                pass
    # At this point the xml data should be fully parsed into objects and object attributes
    # Now we need to match weapons up with their ammo and ammo magazine
    for local_weapon in local_weapon_list:
        for local_ammo in local_ammo_list:
            if local_ammo.name == local_weapon.ammo_magazine_subtype:
                # This ammo object is the one that corresponds with this weapon
                local_weapon.ammo_name = local_ammo.name
                local_weapon.desired_speed = local_ammo.desired_speed
                local_weapon.max_trajectory = local_ammo.max_trajectory
                local_weapon.is_explosive = local_ammo.is_explosive
                local_weapon.backkick_force = local_ammo.backkick_force
                local_weapon.physical_material = local_ammo.physical_material
                local_weapon.projectile_hit_impulse = local_ammo.projectile_hit_impulse
                local_weapon.missile_explosion_radius = local_ammo.missile_explosion_radius
                local_weapon.projectile_mass_damage = local_ammo.projectile_mass_damage
                local_weapon.missile_explosion_damage = local_ammo.missile_explosion_damage
                local_weapon.projectile_health_damage = local_ammo.projectile_health_damage
                local_weapon.missile_health_damage = local_ammo.missile_health_damage
        for local_ammo_magazine in local_ammo_magazine_list:
            if local_ammo_magazine == local_weapon.ammo_magazine_subtype:
                # This ammo magazine is the one that corresponds with this weapon
                local_weapon.ammo_magazine_name = local_ammo_magazine.ammo_magazine_name
                local_weapon.ammo_magazine_capacity = local_ammo_magazine.capacity
    # Here the weapon objects should be fully filled out and are ready to be transcribed into excel
    #print(local_weapon_list)
    excel_setup('main', local_weapon_list)


def update_attr(local_object, local_attribute, local_value):
    if local_attribute == "<SubtypeId>":
        local_object.name = local_value
    elif local_attribute == "RateOfFire=":
        local_object.rate_of_fire = local_value
    elif local_attribute == "ShotsInBurst=":
        local_object.shots_in_burst = local_value
    elif local_attribute == "<ReleaseTimeAfterFire>":
        local_object.release_time_after_fire = local_value
    elif local_attribute == "<DeviateShotAngle>":
        local_object.deviate_shot_angle = local_value
    elif local_attribute == "<AmmoMagazine Subtype=":
        local_object.ammo_magazine_subtype = local_value
    elif local_attribute == "<ReloadTime>":
        local_object.reload_time = local_value
    elif local_attribute == "<DesiredSpeed>":
        local_object.desired_speed = local_value
    elif local_attribute == "<MaxTrajectory>":
        local_object.max_trajectory = local_value
    elif local_attribute == "<IsExplosive>":
        local_object.is_explosive = local_value
    elif local_attribute == "<BackkickForce>":
        local_object.backkick_force = local_value
    elif local_attribute == "<PhysicalMaterial>":
        local_object.physical_material = local_value
    elif local_attribute == "<ProjectileHitImpulse>":
        local_object.projectile_hit_impulse = local_value
    elif local_attribute == "<MissileExplosionRadius>":
        local_object.missile_explosion_radius = local_value
    elif local_attribute == "<ProjectileMassDamage>":
        local_object.projectile_mass_damage = local_value
    elif local_attribute == "<MissileExplosionDamage>":
        local_object.missile_explosion_damage = local_value
    elif local_attribute == "<ProjectileHealthDamage>":
        local_object.projectile_health_damage = local_value
    elif local_attribute == "<MissileHealthDamage>":
        local_object.missile_health_damage = local_value
    elif local_attribute == "<Capacity>":
        local_object.capacity = local_value


# Used to extract a term from an xml line
def remove_term(string, delimiter):
    local_phrase_1 = string.split(delimiter)
    if '=' in delimiter:
        local_finished_delimiter = ' '
        local_phrase_2 = local_phrase_1[1].split(local_finished_delimiter)
        local_phrase_3 = local_phrase_2[0]
        local_phrase_4 = local_phrase_3.translate({ord('"'): None})
        return local_phrase_4
    else:
        local_new_delimiter = delimiter.split('<')
        local_finished_delimiter = local_new_delimiter[1]
        local_phrase_2 = local_phrase_1[1].split(('</' + local_finished_delimiter))
        local_phrase_3 = local_phrase_2[0]
        local_phrase_4 = local_phrase_3.translate({ord('"'): None})
        return local_phrase_4


def excel_setup(file_name, local_weapon_list):
    # Import the Workbook class
    very_local_row = 1
    try:
        local_workbook = openpyxl.load_workbook('modDataParserSEWorkbook.xlsx')
        print("Workbook found existing already")
        local_worksheet = local_workbook.active
        for row in local_worksheet.iter_rows(2, 150, 1, 1, True):
            very_local_row += 1
            if row == (None,):
                local_current_row = very_local_row
                print("Identified first empty row")
                break
            else:
                pass
        for local_completed_weapon in local_weapon_list:
            local_worksheet[('A{}'.format(local_current_row))] = local_completed_weapon.name
            local_worksheet[('B{}'.format(local_current_row))] = local_completed_weapon.rate_of_fire
            local_worksheet[('C{}'.format(local_current_row))] = local_completed_weapon.shots_in_burst
            local_worksheet[('D{}'.format(local_current_row))] = local_completed_weapon.release_time_after_fire
            local_worksheet[('E{}'.format(local_current_row))] = local_completed_weapon.deviate_shot_angle
            local_worksheet[('F{}'.format(local_current_row))] = local_completed_weapon.reload_time
            local_worksheet[('G{}'.format(local_current_row))] = local_completed_weapon.ammo_magazine_subtype
            local_worksheet[('H{}'.format(local_current_row))] = local_completed_weapon.desired_speed
            local_worksheet[('I{}'.format(local_current_row))] = local_completed_weapon.max_trajectory
            local_worksheet[('J{}'.format(local_current_row))] = local_completed_weapon.is_explosive
            local_worksheet[('K{}'.format(local_current_row))] = local_completed_weapon.backkick_force
            local_worksheet[('L{}'.format(local_current_row))] = local_completed_weapon.physical_material
            local_worksheet[('M{}'.format(local_current_row))] = local_completed_weapon.projectile_hit_impulse
            local_worksheet[('N{}'.format(local_current_row))] = local_completed_weapon.missile_explosion_radius
            if local_completed_weapon.physical_material == "GunBullet":
                local_worksheet[('O{}'.format(local_current_row))] = local_completed_weapon.projectile_mass_damage
                local_worksheet[('P{}'.format(local_current_row))] = local_completed_weapon.projectile_health_damage
            else:
                local_worksheet[('O{}'.format(local_current_row))] = local_completed_weapon.missile_explosion_damage
                local_worksheet[('P{}'.format(local_current_row))] = local_completed_weapon.missile_health_damage
            local_worksheet[('Q{}'.format(local_current_row))] = local_completed_weapon.ammo_magazine_capacity
            local_worksheet[('R{}'.format(local_current_row))] = file_name
            local_current_row += 1
    except IOError:
        local_workbook = openpyxl.Workbook()
        local_worksheet = local_workbook.active
        local_worksheet.sheet_properties.tabColor = "1072BA"
        local_worksheet.title = 'Main'
        local_columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']
        # Set up sheet headers
        # Attribute = Weapon.name
        local_worksheet['A1'] = 'Name'
        # Attribute = Weapon.rate_of_fire
        local_worksheet['B1'] = 'Shots/min'
        # Attribute = Weapon.shots_in_burst
        local_worksheet['C1'] = 'Shots/burst'
        # Attribute = Weapon.release_time_after_fire
        local_worksheet['D1'] = 'Cooldown (ms)'
        # Attribute = Weapon.deviate_shot_angle
        local_worksheet['E1'] = 'Inaccuracy (deg)'
        # Attribute = Weapon.reload_time
        local_worksheet['F1'] = 'Reload (ms)'
        # Attribute = Weapon.name
        local_worksheet['G1'] = 'Ammo Name'
        # Attribute = Weapon.desired_speed
        local_worksheet['H1'] = 'Projectile Speed'
        # Attribute = Weapon.max_trajectory
        local_worksheet['I1'] = 'Range'
        # Attribute = Weapon.is_explosive
        local_worksheet['J1'] = 'Explosive?'
        # Attribute = Weapon.backkick_force
        local_worksheet['K1'] = 'Recoil'
        # Attribute = Weapon.physical_material
        local_worksheet['L1'] = 'Type'
        # Attribute = Weapon.projectile_hit_impulse
        local_worksheet['M1'] = 'Hit Impulse'
        # Attribute = Weapon.missile_explosion_radius
        local_worksheet['N1'] = 'Explosion Radius'
        # Attribute = Weapon.projectile_mass_damage OR missile_explosion_damage
        local_worksheet['O1'] = 'General Damage'
        # Attribute = Weapon.projectile_health_damage OR missile_health_damage
        local_worksheet['P1'] = 'Player Damage'
        # Attribute = Weapon.capacity
        local_worksheet['Q1'] = 'Magazine Capacity'
        # Attribute = file_name
        local_worksheet['R1'] = 'Src File Name'
        # Set the column widths
        for column in local_columns:
            local_worksheet.column_dimensions[column].width = 17
        # Set the headers styles
            local_cell = local_worksheet['B2']
            # Freeze row 1, column A
            local_worksheet.freeze_panes = local_cell
            very_local_row = 2
            local_current_row = very_local_row
        for local_completed_weapon in local_weapon_list:
            local_worksheet[('A{}'.format(local_current_row))] = local_completed_weapon.name
            local_worksheet[('B{}'.format(local_current_row))] = local_completed_weapon.rate_of_fire
            local_worksheet[('C{}'.format(local_current_row))] = local_completed_weapon.shots_in_burst
            local_worksheet[('D{}'.format(local_current_row))] = local_completed_weapon.release_time_after_fire
            local_worksheet[('E{}'.format(local_current_row))] = local_completed_weapon.deviate_shot_angle
            local_worksheet[('F{}'.format(local_current_row))] = local_completed_weapon.reload_time
            local_worksheet[('G{}'.format(local_current_row))] = local_completed_weapon.ammo_magazine_subtype
            local_worksheet[('H{}'.format(local_current_row))] = local_completed_weapon.desired_speed
            local_worksheet[('I{}'.format(local_current_row))] = local_completed_weapon.max_trajectory
            local_worksheet[('J{}'.format(local_current_row))] = local_completed_weapon.is_explosive
            local_worksheet[('K{}'.format(local_current_row))] = local_completed_weapon.backkick_force
            local_worksheet[('L{}'.format(local_current_row))] = local_completed_weapon.physical_material
            local_worksheet[('M{}'.format(local_current_row))] = local_completed_weapon.projectile_hit_impulse
            local_worksheet[('N{}'.format(local_current_row))] = local_completed_weapon.missile_explosion_radius
            if local_completed_weapon.physical_material == "GunBullet":
                local_worksheet[('O{}'.format(local_current_row))] = local_completed_weapon.projectile_mass_damage
                local_worksheet[('P{}'.format(local_current_row))] = local_completed_weapon.projectile_health_damage
            else:
                local_worksheet[('O{}'.format(local_current_row))] = local_completed_weapon.missile_explosion_damage
                local_worksheet[('P{}'.format(local_current_row))] = local_completed_weapon.missile_health_damage
            local_worksheet[('Q{}'.format(local_current_row))] = local_completed_weapon.ammo_magazine_capacity
            local_current_row += 1
    local_workbook.save('modDataParserSEWorkbook.xlsx')
    return

main()
input()
